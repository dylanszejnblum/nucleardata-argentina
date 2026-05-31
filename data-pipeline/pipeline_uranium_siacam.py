#!/usr/bin/env python3
"""Pipeline: Ingest uranium data from SIACAM (datos.gob.ar) + CNEA pages.

Structured data ingestion pipeline for uranium:
  - SIACAM mining projects portfolio (XLSX) → 21 uranium projects
  - SIACAM international mineral prices (XLSX) → monthly U price 1990–present
  - SIACAM mineral trade (CSV) → uranium imports/exports
  - CNEA narrative pages → contextual facts

Usage:
    uv run python3 pipeline_uranium_siacam.py --out out_uranium_siacam/
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import httpx
import openpyxl

# ── Sources ──────────────────────────────────────────────────────────────────

SIACAM_PROYECTOS_URL = (
    "https://www.mecon.gob.ar/dataset/"
    "Cartera-de-Proyectos-Mineros-Metaliferos-y-Litio-del-SIACAM.xlsx"
)

SIACAM_PRECIOS_URL = (
    "https://www.mecon.gob.ar/dataset/precios-internacionales-minerales.xlsx"
)

SIACAM_IMPORT_URL = (
    "https://www.mecon.gob.ar/dataset/Comercio/impominerales-siacam.csv"
)

SIACAM_EXPORT_URL = (
    "https://www.mecon.gob.ar/dataset/Comercio/expominerales-siacam.csv"
)

CNEA_PAGES = {
    "que_es": "https://www.argentina.gob.ar/cnea/tecnologia-nuclear/mineria-del-uranio/que-es-el-uranio",
    "exploracion": "https://www.argentina.gob.ar/cnea/tecnologia-nuclear/mineria-del-uranio/exploracion-de-uranio",
    "produccion": "https://www.argentina.gob.ar/cnea/tecnologia-nuclear/mineria-del-uranio/produccion-de-uranio",
    "remediacion": "https://www.argentina.gob.ar/cnea/tecnologia-nuclear/mineria-del-uranio/remediacion-ambiental-de-la-mineria-del-uranio",
}

STATUS_ORDER = {
    "Prospección": 0,
    "Exploración inicial": 1,
    "Exploración avanzada": 2,
    "Evaluación Económica Preliminar": 3,
    "Factibilidad": 4,
}

STATUS_EN = {
    "Prospección": "prospection",
    "Exploración inicial": "early_exploration",
    "Exploración avanzada": "advanced_exploration",
    "Evaluación Económica Preliminar": "preliminary_economic_assessment",
    "Factibilidad": "feasibility",
}

PROVINCE_CODES = {
    "Río Negro": "RN",
    "Chubut": "CT",
    "Neuquén": "NQ",
    "Mendoza": "MZ",
    "Salta": "SA",
    "Santa Cruz": "SC",
}


# ── Helpers ──────────────────────────────────────────────────────────────────


def download(url: str, dest: Path) -> None:
    """Download a file to dest if it doesn't already exist."""
    if dest.exists():
        print(f"  ✓ Already cached: {dest.name}")
        return
    print(f"  ↓ Downloading {dest.name}...")
    import subprocess, shlex
    cmd = ["curl", "-sSL", "--connect-timeout", "15", "--max-time", "60", url, "-o", str(dest)]
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0:
        print(f"  ✗ curl failed (exit {r.returncode}): {r.stderr[:200]}", file=sys.stderr)
        raise SystemExit(1)
    print(f"  ✓ Saved {dest.stat().st_size:,} bytes → {dest.name}")


def clean_text(html: str) -> str:
    """Strip HTML tags and normalize whitespace."""
    text = re.sub(r"<script[^>]*>.*?</script>", "", html, flags=re.DOTALL)
    text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


# ── Stage 1: Raw Snapshots ──────────────────────────────────────────────────


def stage1_raw(out_dir: Path) -> dict:
    """Download all source files."""
    raw_dir = out_dir / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    checkpoint = raw_dir / "downloads_complete.json"

    if checkpoint.exists():
        print("  [stage1] checkpoint exists — skipping download")
        return json.loads(checkpoint.read_text())

    manifest = {"downloads": [], "timestamp": datetime.utcnow().isoformat()}

    # SIACAM XLSX files
    for url, name in [
        (SIACAM_PROYECTOS_URL, "siacam_proyectos.xlsx"),
        (SIACAM_PRECIOS_URL, "siacam_precios.xlsx"),
        (SIACAM_IMPORT_URL, "siacam_importaciones.csv"),
        (SIACAM_EXPORT_URL, "siacam_exportaciones.csv"),
    ]:
        dest = raw_dir / name
        download(url, dest)
        manifest["downloads"].append({"name": name, "size": dest.stat().st_size})

    # CNEA pages
    for key, url in CNEA_PAGES.items():
        dest = raw_dir / f"cnea_{key}.html"
        if not dest.exists():
            print(f"  ↓ Downloading cnea_{key}.html...")
            import subprocess
            cmd = ["curl", "-sSL", "--connect-timeout", "15", "--max-time", "30", url, "-o", str(dest)]
            r = subprocess.run(cmd, capture_output=True, text=True)
            if r.returncode != 0:
                print(f"  ✗ curl failed: {r.stderr[:200]}", file=sys.stderr)
                raise SystemExit(1)
        manifest["downloads"].append(
            {"name": f"cnea_{key}.html", "size": dest.stat().st_size}
        )

    checkpoint.write_text(json.dumps(manifest, indent=2, ensure_ascii=False))
    print(f"  [stage1] {len(manifest['downloads'])} files downloaded")
    return manifest


# ── Stage 2: Staging — Extract uranium data ─────────────────────────────────


def stage2_stage(out_dir: Path) -> dict:
    """Extract uranium rows from raw files into staging CSVs."""
    raw_dir = out_dir / "raw"
    stage_dir = out_dir / "staging"
    stage_dir.mkdir(parents=True, exist_ok=True)
    checkpoint = stage_dir / "stage_complete.json"

    if checkpoint.exists():
        print("  [stage2] checkpoint exists — skipping")
        return json.loads(checkpoint.read_text())

    manifest = {"timestamp": datetime.utcnow().isoformat(), "files": []}

    # ── Projects ─────────────────────────────────────────────────────────
    proyectos_path = raw_dir / "siacam_proyectos.xlsx"
    wb = openpyxl.load_workbook(proyectos_path)
    ws = wb.active
    headers = [str(ws.cell(row=1, column=c).value or "") for c in range(1, 17)]
    COL_NOMBRE = 1
    COL_LAT = 2
    COL_LNG = 3
    COL_MINERAL = 4
    COL_PROV = 5
    COL_STATUS = 6
    COL_CTRL1 = 7
    COL_PCT1 = 8
    COL_ORIG1 = 9
    COL_CTRL2 = 10
    COL_ORIG2 = 11
    COL_PCT2 = 12

    # Find column indices by header name
    for i, h in enumerate(headers):
        h_clean = h.strip().lower().replace("°", "").replace(")", "").replace("(", "")
        if "nombre" in h_clean:
            COL_NOMBRE = i
        elif "latitud" in h_clean:
            COL_LAT = i
        elif "longitud" in h_clean:
            COL_LNG = i
        elif "mineral" in h_clean:
            COL_MINERAL = i
        elif "provincia" in h_clean:
            COL_PROV = i
        elif "estado" in h_clean:
            COL_STATUS = i
        elif "controlante" in h_clean and "1" in h_clean or i == 6:
            COL_CTRL1 = i
        elif "porcentaje" in h_clean and "1" in h_clean or i == 7:
            COL_PCT1 = i
        elif "origen" in h_clean and "1" in h_clean or i == 8:
            COL_ORIG1 = i

    projects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = [str(v) if v is not None else "" for v in row]
        if not vals[COL_MINERAL]:
            continue
        mineral = vals[COL_MINERAL].strip().lower()
        if "uran" not in mineral:
            continue

        lat_str = vals[COL_LAT].strip()
        lng_str = vals[COL_LNG].strip()
        lat = float(lat_str) if lat_str else None
        lng = float(lng_str) if lng_str else None
        status_raw = vals[COL_STATUS].strip() if len(vals) > COL_STATUS else ""
        status_label = status_raw if status_raw in STATUS_ORDER else "Unknown"

        ctrl1 = vals[COL_CTRL1].strip() if len(vals) > COL_CTRL1 else ""
        pct1 = vals[COL_PCT1].strip() if len(vals) > COL_PCT1 else ""
        orig1 = vals[COL_ORIG1].strip() if len(vals) > COL_ORIG1 else ""

        projects.append(
            {
                "project_name": vals[COL_NOMBRE].strip(),
                "latitude": lat,
                "longitude": lng,
                "province": vals[COL_PROV].strip(),
                "province_code": PROVINCE_CODES.get(vals[COL_PROV].strip(), ""),
                "status": STATUS_EN.get(status_label, status_label.lower()),
                "status_label": status_label,
                "mineral": "Uranio",
                "controllers": [
                    {
                        "name": ctrl1 or "-",
                        "ownership_pct": pct1,
                        "origin_country": orig1 or "-",
                    }
                ],
            }
        )

    projects_path_out = stage_dir / "uranium_projects.csv"
    with open(projects_path_out, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=projects[0].keys())
        w.writeheader()
        w.writerows(projects)
    print(f"  [stage2] {len(projects)} uranium projects → {projects_path_out.name}")
    manifest["files"].append(
        {"name": projects_path_out.name, "rows": len(projects)}
    )

    # ── Prices ───────────────────────────────────────────────────────────
    precios_path = raw_dir / "siacam_precios.xlsx"
    wb2 = openpyxl.load_workbook(precios_path)
    ws2 = wb2["Sheet1"]

    prices = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        mineral_raw = str(row[4] or "").strip().lower()
        if "uran" not in mineral_raw:
            continue
        year = row[0]
        month = row[1]
        price = float(row[3]) if row[3] is not None else None
        unit = str(row[5] or "")
        fecha = f"{int(year)}-{int(month):02d}-01"
        prices.append(
            {
                "date": fecha,
                "year": int(year),
                "month": int(month),
                "price_usd": price,
                "unit": unit,
                "mineral": "Uranio (U)",
            }
        )

    prices_out = stage_dir / "uranium_prices.csv"
    with open(prices_out, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=prices[0].keys())
        w.writeheader()
        w.writerows(prices)
    print(f"  [stage2] {len(prices)} price rows → {prices_out.name}")
    manifest["files"].append({"name": prices_out.name, "rows": len(prices)})

    # ── Trade: imports ───────────────────────────────────────────────────
    imp_path = raw_dir / "siacam_importaciones.csv"
    imports = []
    with open(imp_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            grupo = (row.get("GRUPO", "") or "").strip().lower()
            if "uran" not in grupo:
                continue
            imports.append(
                {
                    "year": int(row.get("ANYO", 0)),
                    "month": int(row.get("MES", 1)),
                    "country": row.get("ORIGEN", "").strip(),
                    "procedence": row.get("PROCEDENC", "").strip(),
                    "value_cif_usd": float(row.get("CIF", 0) or 0),
                    "trade_type": "import",
                    "sector": row.get("SECTOR", ""),
                }
            )
    imp_out = stage_dir / "uranium_imports.csv"
    if imports:
        with open(imp_out, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=imports[0].keys())
            w.writeheader()
            w.writerows(imports)
    print(f"  [stage2] {len(imports)} import rows → {imp_out.name}")
    manifest["files"].append({"name": imp_out.name, "rows": len(imports)})

    # ── Trade: exports ───────────────────────────────────────────────────
    exp_path = raw_dir / "siacam_exportaciones.csv"
    exports = []
    with open(exp_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            grupo = (row.get("GRUPO", "") or "").strip().lower()
            if "uran" not in grupo:
                continue
            exports.append(
                {
                    "year": int(row.get("ANYO", 0)),
                    "month": int(row.get("MES", 1)),
                    "country": row.get("PAIS", row.get("ORIGEN", "")).strip(),
                    "value_fob_usd": float(row.get("FOB", 0) or 0),
                    "trade_type": "export",
                    "sector": row.get("SECTOR", ""),
                }
            )
    exp_out = stage_dir / "uranium_exports.csv"
    if exports:
        with open(exp_out, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=exports[0].keys())
            w.writeheader()
            w.writerows(exports)
    print(f"  [stage2] {len(exports)} export rows → {exp_out.name}")
    manifest["files"].append({"name": exp_out.name, "rows": len(exports)})

    # ── CNEA facts ───────────────────────────────────────────────────────
    facts = {
        "total_resources_tU": 33650,
        "historical_production_tU": 2600,
        "production_years": "1952-1997",
        "last_production_year": 1997,
        "first_deposit_year": 1946,
        "first_deposit_location": "Las Heras, Mendoza",
        "first_yellowcake_year": 1952,
        "first_yellowcake_location": "Córdoba",
        "first_yellowcake_amount_t": 4.5,
        "identified_deposits": 13,
        "production_centers": 7,
        "remediation_sites": 7,
        "remediation_provinces": [
            "Mendoza", "Córdoba", "San Luis", "La Rioja", "Salta", "Chubut"
        ],
        "producing_provinces": [
            "Río Negro", "Chubut", "Neuquén", "Mendoza", "Salta", "Santa Cruz"
        ],
        "active_projects_count": 21,
    }
    facts_out = stage_dir / "cnea_facts.json"
    facts_out.write_text(json.dumps(facts, indent=2, ensure_ascii=False))
    print(f"  [stage2] CNEA facts → {facts_out.name}")
    manifest["files"].append({"name": facts_out.name, "rows": 1})

    checkpoint.write_text(json.dumps(manifest, indent=2, ensure_ascii=False))
    return manifest


# ── Stage 3: Normalized output ──────────────────────────────────────────────


def stage3_normalize(out_dir: Path) -> dict:
    """Normalize to JSON output files for backend consumption."""
    stage_dir = out_dir / "staging"
    norm_dir = out_dir / "normalized"
    norm_dir.mkdir(parents=True, exist_ok=True)
    checkpoint = norm_dir / "normalize_complete.json"

    if checkpoint.exists():
        print("  [stage3] checkpoint exists — skipping")
        return json.loads(checkpoint.read_text())

    manifest = {"timestamp": datetime.utcnow().isoformat(), "files": []}

    # Projects
    projects_list = []
    with open(stage_dir / "uranium_projects.csv", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            projects_list.append(row)

    norm_projects = norm_dir / "uranium_projects.json"
    norm_projects.write_text(
        json.dumps(
            {
                "count": len(projects_list),
                "mineral": "Uranio",
                "projects": projects_list,
                "generated_at": datetime.utcnow().isoformat(),
            },
            indent=2,
            ensure_ascii=False,
        )
    )
    manifest["files"].append(
        {"name": "uranium_projects.json", "rows": len(projects_list)}
    )

    # Prices
    prices_list = []
    with open(stage_dir / "uranium_prices.csv", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            row["price_usd"] = float(row["price_usd"])
            prices_list.append(row)

    prices_vals = [p["price_usd"] for p in prices_list if p["price_usd"]]
    norm_prices = norm_dir / "uranium_prices.json"
    norm_prices.write_text(
        json.dumps(
            {
                "count": len(prices_list),
                "mineral": "Uranio",
                "unit": "USD/lb",
                "source": "SIACAM (FRED → Cameco)",
                "prices": prices_list,
                "stats": {
                    "min": min(prices_vals),
                    "min_date": next(
                        p["date"] for p in prices_list if p["price_usd"] == min(prices_vals)
                    ),
                    "max": max(prices_vals),
                    "max_date": next(
                        p["date"] for p in prices_list if p["price_usd"] == max(prices_vals)
                    ),
                    "current": prices_vals[-1] if prices_vals else None,
                    "current_date": prices_list[-1]["date"] if prices_list else None,
                    "all_time_high": max(prices_vals),
                    "all_time_low": min(prices_vals),
                },
                "generated_at": datetime.utcnow().isoformat(),
            },
            indent=2,
            ensure_ascii=False,
        )
    )
    manifest["files"].append(
        {"name": "uranium_prices.json", "rows": len(prices_list)}
    )

    # Trade
    for trade_file, trade_name in [
        ("uranium_imports.csv", "uranium_imports.json"),
        ("uranium_exports.csv", "uranium_exports.json"),
    ]:
        rows = []
        csv_path = stage_dir / trade_file
        if csv_path.exists():
            with open(csv_path, newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    rows.append(row)
        out_path = norm_dir / trade_name
        out_path.write_text(
            json.dumps(
                {
                    "count": len(rows),
                    "mineral": "Uranio",
                    "records": rows,
                    "generated_at": datetime.utcnow().isoformat(),
                },
                indent=2,
                ensure_ascii=False,
            )
        )
        manifest["files"].append({"name": trade_name, "rows": len(rows)})

    # CNEA facts
    facts = json.loads((stage_dir / "cnea_facts.json").read_text())
    norm_facts = norm_dir / "uranium_facts.json"
    norm_facts.write_text(
        json.dumps(
            {
                "mineral": "Uranio",
                "source": "CNEA — argentina.gob.ar",
                "facts": facts,
                "generated_at": datetime.utcnow().isoformat(),
            },
            indent=2,
            ensure_ascii=False,
        )
    )
    manifest["files"].append({"name": "uranium_facts.json", "rows": 1})

    checkpoint.write_text(json.dumps(manifest, indent=2, ensure_ascii=False))
    return manifest


# ── Stage 4: Aggregates ─────────────────────────────────────────────────────


def stage4_aggregate(out_dir: Path) -> dict:
    """Compute and write aggregate statistics."""
    norm_dir = out_dir / "normalized"
    agg_dir = out_dir / "aggregates"
    agg_dir.mkdir(parents=True, exist_ok=True)
    checkpoint = agg_dir / "aggregate_complete.json"

    if checkpoint.exists():
        print("  [stage4] checkpoint exists — skipping")
        return json.loads(checkpoint.read_text())

    manifest = {"timestamp": datetime.utcnow().isoformat(), "files": []}

    # Load projects
    projects = json.loads((norm_dir / "uranium_projects.json").read_text())[
        "projects"
    ]

    # By status
    by_status = {}
    for p in projects:
        s = p.get("status_label", "Unknown")
        by_status[s] = by_status.get(s, 0) + 1

    # By province
    by_province = {}
    for p in projects:
        prov = p.get("province", "Unknown")
        by_province[prov] = by_province.get(prov, 0) + 1

    # By controller origin
    by_origin = {}
    for p in projects:
        ctrls = p.get("controllers", [])
        if ctrls and isinstance(ctrls, list):
            origin = ctrls[0].get("origin_country", "-")
        else:
            origin = "-"
        by_origin[origin] = by_origin.get(origin, 0) + 1

    # By controller company
    by_company = {}
    for p in projects:
        ctrls = p.get("controllers", [])
        if ctrls and isinstance(ctrls, list):
            company = ctrls[0].get("name", "-")
        else:
            company = "-"
        by_company[company] = by_company.get(company, 0) + 1

    agg = {
        "total_projects": len(projects),
        "projects_by_status": by_status,
        "projects_by_province": by_province,
        "projects_by_origin": by_origin,
        "projects_by_company": by_company,
    }

    agg_path = agg_dir / "project_aggregates.json"
    agg_path.write_text(
        json.dumps(agg, indent=2, ensure_ascii=False)
    )
    manifest["files"].append({"name": "project_aggregates.json", "rows": 1})

    # Price statistics
    prices = json.loads((norm_dir / "uranium_prices.json").read_text())
    price_stats = {
        "total_data_points": prices["count"],
        "unit": prices["unit"],
        "source": prices["source"],
        "current_price_usd": prices["stats"]["current"],
        "current_date": prices["stats"]["current_date"],
        "all_time_high_usd": prices["stats"]["all_time_high"],
        "all_time_high_date": prices["stats"]["max_date"],
        "all_time_low_usd": prices["stats"]["all_time_low"],
        "all_time_low_date": prices["stats"]["min_date"],
    }

    # Decade averages
    decade_avgs = {}
    for p in prices["prices"]:
        decade = f"{int(p['year']) // 10 * 10}s"
        if decade not in decade_avgs:
            decade_avgs[decade] = {"total": 0, "count": 0}
        decade_avgs[decade]["total"] += p["price_usd"]
        decade_avgs[decade]["count"] += 1
    price_stats["decade_averages"] = {
        k: round(v["total"] / v["count"], 2)
        for k, v in sorted(decade_avgs.items())
    }

    stats_path = agg_dir / "price_statistics.json"
    stats_path.write_text(
        json.dumps(price_stats, indent=2, ensure_ascii=False)
    )
    manifest["files"].append({"name": "price_statistics.json", "rows": 1})

    checkpoint.write_text(json.dumps(manifest, indent=2, ensure_ascii=False))
    return manifest


# ── Stage 5: Latest snapshot for dashboard ──────────────────────────────────


def stage5_latest(out_dir: Path) -> dict:
    """Write out/latest.json — lightweight snapshot for dashboard frontend."""
    norm_dir = out_dir / "normalized"
    agg_dir = out_dir / "aggregates"
    latest_path = out_dir / "latest.json"

    projects = json.loads((norm_dir / "uranium_projects.json").read_text())
    prices = json.loads((norm_dir / "uranium_prices.json").read_text())
    facts = json.loads((norm_dir / "uranium_facts.json").read_text())
    agg = json.loads((agg_dir / "project_aggregates.json").read_text())
    stats = json.loads((agg_dir / "price_statistics.json").read_text())

    latest = {
        "mineral": "Uranio",
        "current_price_usd": prices["stats"]["current"],
        "current_date": prices["stats"]["current_date"],
        "price_change_pct": round(
            (prices["prices"][-1]["price_usd"] / prices["prices"][-2]["price_usd"] - 1)
            * 100,
            2,
        )
        if len(prices["prices"]) >= 2
        else 0,
        "total_resources_tU": facts["facts"]["total_resources_tU"],
        "historical_production_tU": facts["facts"]["historical_production_tU"],
        "active_projects": agg["total_projects"],
        "provinces_with_projects": len(agg["projects_by_province"]),
        "advanced_projects": sum(
            v
            for k, v in agg["projects_by_status"].items()
            if k
            in ("Exploración avanzada", "Evaluación Económica Preliminar", "Factibilidad")
        ),
        "projects_by_status": agg["projects_by_status"],
        "projects_by_province": agg["projects_by_province"],
        "price_stats": stats,
        "generated_at": datetime.utcnow().isoformat(),
    }

    latest_path.write_text(json.dumps(latest, indent=2, ensure_ascii=False))
    print(f"  [stage5] → {latest_path}")
    return {"file": str(latest_path)}


# ── Main ─────────────────────────────────────────────────────────────────────


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Uranium SIACAM + CNEA data pipeline"
    )
    ap.add_argument(
        "--out",
        default="out_uranium_siacam",
        type=Path,
        help="Output directory (default: out_uranium_siacam)",
    )
    ap.add_argument(
        "--skip-download",
        action="store_true",
        help="Skip download stage (use cached raw files)",
    )
    args = ap.parse_args()

    t0 = time.time()

    print("\n═══════════════════════════════════════════════")
    print("  URANIO — SIACAM + CNEA Data Pipeline")
    print("═══════════════════════════════════════════════\n")

    # Stage 1: Raw snapshots
    print("┌─ Stage 1: Raw snapshots")
    if args.skip_download:
        print("  └ (skipped via --skip-download)")
    else:
        stage1_raw(args.out)
    print()

    # Stage 2: Extraction & Staging
    print("┌─ Stage 2: Extraction & Staging")
    stage2_stage(args.out)
    print()

    # Stage 3: Normalization
    print("┌─ Stage 3: Normalization")
    stage3_normalize(args.out)
    print()

    # Stage 4: Aggregates
    print("┌─ Stage 4: Aggregates")
    stage4_aggregate(args.out)
    print()

    # Stage 5: Latest snapshot
    print("┌─ Stage 5: Latest snapshot")
    stage5_latest(args.out)
    print()

    elapsed = time.time() - t0
    print(f"✅ Pipeline complete in {elapsed:.1f}s")
    print(f"   Output: {args.out.resolve()}")
    print(f"   ├── raw/          — original downloads")
    print(f"   ├── staging/      — extracted/filtered CSV")
    print(f"   ├── normalized/   — JSON for backend seed")
    print(f"   ├── aggregates/   — computed summaries")
    print(f"   └── latest.json   — dashboard-ready snapshot")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
